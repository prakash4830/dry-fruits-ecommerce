"""
Report generation service for Nectar & Nut E-commerce.

Worker: Dev - PDF and Excel report generation
Worker: BA - Maps to FR-R01 through FR-R05
Worker: QA - Unit tests required
"""

import io
from datetime import datetime, timedelta
from decimal import Decimal

import pandas as pd
from django.db.models import Sum, Count, Avg
from django.utils import timezone

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from orders.models import Order, OrderItem
from products.models import Product


class ReportGenerator:
    """
    Report generation service using Pandas, ReportLab, and Openpyxl.
    
    Worker: Dev - Report generation logic
    Worker: QA - Validate report accuracy
    """
    
    def __init__(self):
        self.company_name = "Nectar & Nut"
        self.company_tagline = "Premium Dry Fruits & Candies"
    
    def get_sales_data(self, start_date, end_date):
        """
        Query sales data for date range.
        
        Worker: Dev - Data aggregation using Pandas
        """
        orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            payment_status='paid'
        ).order_by('created_at')
        
        # Convert to DataFrame
        data = []
        for order in orders:
            for item in order.items.all():
                data.append({
                    'date': order.created_at.date(),
                    'order_number': order.order_number,
                    'product': item.product_name,
                    'quantity': item.quantity,
                    'unit_price': float(item.price),
                    'total': float(item.total),
                    'status': order.status,
                })
        
        if data:
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame(columns=[
                'date', 'order_number', 'product', 'quantity',
                'unit_price', 'total', 'status'
            ])
        
        return df, orders
    
    def get_summary_stats(self, orders):
        """
        Calculate summary statistics.
        
        Worker: Dev - Aggregation for report headers
        """
        if not orders.exists():
            return {
                'total_orders': 0,
                'total_revenue': Decimal('0.00'),
                'total_items': 0,
                'avg_order_value': Decimal('0.00'),
            }
        
        stats = orders.aggregate(
            total_orders=Count('id'),
            total_revenue=Sum('total_amount'),
            avg_order_value=Avg('total_amount'),
        )
        
        total_items = OrderItem.objects.filter(
            order__in=orders
        ).aggregate(total=Sum('quantity'))['total'] or 0
        
        return {
            'total_orders': stats['total_orders'] or 0,
            'total_revenue': stats['total_revenue'] or Decimal('0.00'),
            'total_items': total_items,
            'avg_order_value': stats['avg_order_value'] or Decimal('0.00'),
        }
    
    def generate_sales_pdf(self, start_date, end_date):
        """
        Generate sales report as PDF.
        
        Worker: Dev - PDF generation using ReportLab
        Worker: BA - Maps to FR-R01
        """
        df, orders = self.get_sales_data(start_date, end_date)
        stats = self.get_summary_stats(orders)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=24,
            alignment=1,  # Center
            spaceAfter=20
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=1,
            textColor=colors.grey
        )
        
        elements = []
        
        # Header
        elements.append(Paragraph(self.company_name, title_style))
        elements.append(Paragraph(self.company_tagline, subtitle_style))
        elements.append(Spacer(1, 20))
        
        # Report Title
        elements.append(Paragraph(
            f"Sales Report: {start_date} to {end_date}",
            styles['Heading2']
        ))
        elements.append(Spacer(1, 20))
        
        # Summary Table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Orders', str(stats['total_orders'])],
            ['Total Revenue', f"₹{stats['total_revenue']:,.2f}"],
            ['Total Items Sold', str(stats['total_items'])],
            ['Average Order Value', f"₹{stats['avg_order_value']:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Orders Table
        if not df.empty:
            elements.append(Paragraph("Order Details", styles['Heading3']))
            elements.append(Spacer(1, 10))
            
            # Prepare table data
            table_data = [['Date', 'Order #', 'Product', 'Qty', 'Price', 'Total']]
            for _, row in df.iterrows():
                table_data.append([
                    str(row['date']),
                    row['order_number'][:12],
                    row['product'][:20],
                    str(row['quantity']),
                    f"₹{row['unit_price']:.2f}",
                    f"₹{row['total']:.2f}",
                ])
            
            orders_table = Table(table_data, colWidths=[
                0.9*inch, 1.2*inch, 1.8*inch, 0.5*inch, 0.9*inch, 0.9*inch
            ])
            orders_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            
            elements.append(orders_table)
        else:
            elements.append(Paragraph(
                "No orders found for the selected date range.",
                styles['Normal']
            ))
        
        # Footer
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}",
            subtitle_style
        ))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_sales_excel(self, start_date, end_date):
        """
        Generate sales report as Excel.
        
        Worker: Dev - Excel generation using Openpyxl
        Worker: BA - Maps to FR-R01
        """
        df, orders = self.get_sales_data(start_date, end_date)
        stats = self.get_summary_stats(orders)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = self.company_name
        ws['A1'].font = Font(bold=True, size=20)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Sales Report: {start_date} to {end_date}"
        ws['A2'].font = Font(size=14)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Summary Section
        ws['A4'] = "Summary"
        ws['A4'].font = Font(bold=True, size=14)
        
        summary_items = [
            ('Total Orders:', stats['total_orders']),
            ('Total Revenue:', f"₹{stats['total_revenue']:,.2f}"),
            ('Total Items Sold:', stats['total_items']),
            ('Average Order Value:', f"₹{stats['avg_order_value']:,.2f}"),
        ]
        
        for i, (label, value) in enumerate(summary_items, start=5):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
        
        # Order Details
        start_row = 11
        ws[f'A{start_row}'] = "Order Details"
        ws[f'A{start_row}'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ['Date', 'Order Number', 'Product', 'Quantity', 'Unit Price', 'Total']
        header_row = start_row + 1
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        if not df.empty:
            for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
                ws.cell(row=row_idx, column=1, value=str(row['date'])).border = border
                ws.cell(row=row_idx, column=2, value=row['order_number']).border = border
                ws.cell(row=row_idx, column=3, value=row['product']).border = border
                ws.cell(row=row_idx, column=4, value=row['quantity']).border = border
                ws.cell(row=row_idx, column=5, value=f"₹{row['unit_price']:.2f}").border = border
                ws.cell(row=row_idx, column=6, value=f"₹{row['total']:.2f}").border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_inventory_excel(self):
        """
        Generate inventory report as Excel.
        
        Worker: Dev - Inventory report
        Worker: BA - Maps to FR-R03
        """
        products = Product.objects.all().order_by('category__name', 'name')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        low_stock_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{self.company_name} - Inventory Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['SKU', 'Product Name', 'Category', 'Weight', 'Price', 'Stock', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, product in enumerate(products, start=5):
            ws.cell(row=row_idx, column=1, value=product.sku or '-')
            ws.cell(row=row_idx, column=2, value=product.name)
            ws.cell(row=row_idx, column=3, value=product.category.name)
            ws.cell(row=row_idx, column=4, value=product.weight)
            ws.cell(row=row_idx, column=5, value=f"₹{product.price}")
            ws.cell(row=row_idx, column=6, value=product.stock)
            
            if product.stock == 0:
                status = "Out of Stock"
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).fill = low_stock_fill
            elif product.is_low_stock:
                status = "Low Stock"
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
                    )
            else:
                status = "In Stock"
            
            ws.cell(row=row_idx, column=7, value=status)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# Singleton instance
report_generator = ReportGenerator()
